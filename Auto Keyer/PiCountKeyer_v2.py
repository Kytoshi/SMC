import tkinter as tk
import customtkinter as ctk
import openpyxl
from tkinter import *
from tkinter import filedialog
from PIL import Image, ImageTk
import pyautogui
import time
import threading
from pynput import keyboard

##### Program Global Settings #####

ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

global stop_flag, entryVar, labelVar, cellVar, pathVar, sheetVar
stop_flag = False

##### Clear Page Configuration #####
""" Page for the clearing function input """

class ClearPage(ctk.CTkFrame):
    def __init__(self, parent, controller):
        super().__init__(parent)
        self.controller = controller

        global entryVar, labelVar

        entryVar = tk.StringVar()
        labelVar = tk.StringVar()

        self.configure(fg_color="#ECEFCA")

        title2 = ctk.CTkLabel(self,
        text="Clearing Page", 
        font=("Arial", 26, "bold"), 
        text_color="#213448", 
        anchor="center", 
        width=700
        )
        title2.place(relx=0.5, rely=0.2, anchor=CENTER)

        # BACK BUTTON ---

        back_default_img = ImageTk.PhotoImage(Image.open("components/defaultback.png").resize((60, 60), Image.Resampling.BICUBIC))
        
        self.back_button = ctk.CTkButton(self, 
            image=back_default_img, 
            width=40, height=40,
            text="",
            fg_color="transparent",
            hover_color="#ECEFCA",
            command=lambda: self.backhome()
            )
        self.back_button.place(relx=0.1, rely=0.2, anchor=CENTER)

        prompt2 = ctk.CTkLabel(self,
        text="How many boxes do you need to clear?",
        font=("Arial", 15, "bold"),
        text_color="#213448",
        anchor="center",
        width=200
        )
        prompt2.place(relx=0.5, rely=0.3, anchor=CENTER)

        self.clear_entry = ctk.CTkEntry(self, placeholder_text="Enter number of boxes...", width=300, height=40, textvariable=entryVar)
        self.clear_entry.place(relx=0.5, rely=0.45, anchor=CENTER)
        # clear_entry.bind("<Return>", start_clear)

        begin_clearing_button = ctk.CTkButton(self, text="Confirm", font=("Arial", 17, "bold"), width=200, height=40, command=self.validate_input)
        begin_clearing_button.place(relx=0.5, rely=0.65, anchor=CENTER)

    def validate_input(self):
        value = entryVar.get()
        try:
            if int(value) > 0:
                self.clear_entry.configure(border_color="#A9A9A9")  # Reset to default gray
                self.controller.show_page("countDown")
            else:
                raise ValueError
        except ValueError:
            self.clear_entry.configure(border_color="red")  # Show red border
    
    def backhome(self):
            """Return to Home and reset flags."""
            entryVar.set("")  # Clear the entry field
            self.clear_entry.configure(border_color="#A9A9A9")  # Reset to default gray
            self.controller.show_page("Home")  # Go back to home

##### Count Down Page Configuration #####
""" Supporter Page for the Clear Page."""

class countDownPage(ctk.CTkFrame):
    def __init__(self, parent, controller, countdown_time=5):
        super().__init__(parent)
        self.controller = controller
        self.countdown_time = countdown_time
        self._cancelled = False  # Add cancellation flag

        self.configure(fg_color="#ECEFCA")

        self.count_text = ctk.CTkLabel(self, text="Starting in", font=("Arial", 26, "bold"), text_color="#213448")
        self.count_text.place(relx=0.5, rely=0.3, anchor=CENTER)

        self.count = ctk.CTkLabel(self, text="", font=("Arial", 40, "bold"), text_color="#213448")
        self.count.place(relx=0.5, rely=0.4, anchor=CENTER)

        # Home button (hidden at start)
        self.home_button = ctk.CTkButton(
            self,
            text="Return to Home",
            font=("Arial", 17, "bold"),
            width=200,
            height=40,
            command=lambda: self.return_home()  # Use method to return to home
        )
        self.home_button.place(relx=0.5, rely=0.6, anchor=CENTER)
        self.home_button.lower()  # Hide initially

        # Cancel button
        self.cancel_button = ctk.CTkButton(
            self,
            text="Cancel",
            font=("Arial", 17, "bold"),
            width=200,
            height=40,
            command=self.cancel_countdown  # Cancel countdown
        )
        self.cancel_button.place(relx=0.5, rely=0.7, anchor=CENTER)

    def cancel_countdown(self):
        """Cancel countdown and reset the flag."""
        self._cancelled = True  # Set the cancel flag
        global stop_flag
        stop_flag = True  # Set stop_flag to True to stop the process
        self.count.configure(text="Cancelled", text_color="red")
        entryVar.set("")  # Clear the entry field after processing
        self.count_text.configure(text="")  # Clear the "Starting in" text
        self.home_button.lift()  # Show the Home button so user can go back manually
        self.cancel_button.lower()  # Hide the Cancel button after cancellation

    def return_home(self):
        """Return to Home and reset flags."""
        global stop_flag
        stop_flag = False  # Reset stop flag for future runs
        self._cancelled = False  # Reset cancel flag
        self.count_text.configure(text="Starting in", text_color="#213448")
        self.controller.show_page("Home")  # Go back to home

    def start_countdown(self):
        self.controller.pages["Clear"].clear_entry.configure(border_color="#A9A9A9")
        self.remaining = self.countdown_time
        self.cancel_button.lift()
        self.home_button.lower()  # Hide button if revisiting
        self.count.configure(text="", text_color="#213448")
        self._update_timer()

    def _update_timer(self):
        if self.remaining >= 0 and not self._cancelled:  # Check if not cancelled
            self.count.configure(text=self.remaining)
            self.remaining -= 1
            self.update()  # Force UI update to display Cancel button
            self.after(1000, self._update_timer)
        else:
            if not self._cancelled:
                self.count_text.configure(text="")
                self.count.configure(text="Clearing has started!")
                self.after(500, self.on_clear_finished)  # Call on_clear_finished after countdown
            else:
                # If cancelled, stop further action and reset
                self.count.configure(text="Cancelled", text_color="red")

    def on_clear_finished(self):
        """Start the clearing process in a separate thread."""
        thread = threading.Thread(target=self.start_clear_thread)
        thread.daemon = True # Make thread a daemon so it exits when the main program exits
        thread.start()

    def start_clear_thread(self):
        """Runs the clearing logic and updates UI from main thread."""
        self.start_clear()

    def start_clear(self):
        """Converts the entry value into an integer and starts the clearing process; cancels process if stop_flag is set."""
        global entryVar, stop_flag, labelVar
        boxes = entryVar.get()
        try:
            boxes = int(boxes)
        except ValueError:
            self.count.configure(text="Invalid input", text_color="red")
            return

        labelVar.set(boxes)
        entryVar.set("")  # Clear the entry field after processing

        for count in range(boxes):
            if stop_flag:
                print("Process stopped by user.")
                break
            self.clear_boxes(-1)
            time.sleep(0.1)  # Optional: Adjust time delay between inputs if needed

        if not stop_flag:
            self.after(1000, self._update_done)  # Update UI after process is done
        else:
            self.after(1000, self._update_cancelled)  # Update UI after process is cancelled

    def _update_done(self):
        """Update the UI after the clearing process is done."""
        self.count.configure(text="Done!", text_color="green")
        self.cancel_button.lower()  # Hide the cancel button after finishing
        self.home_button.lift()  # Show the home button after finishing

    def _update_cancelled(self):
        """Update the UI after the clearing process is cancelled."""
        self.count.configure(text="Cancelled", text_color="red")
        self.home_button.lift()  # Show the home button after cancelling

    def clear_boxes(self, data):
        """Deleting Count in Input."""
        # Type the data as a string
        pyautogui.typewrite(str(data))
        time.sleep(0.1)
        pyautogui.press("enter")  # Press Enter to press confirm button
        time.sleep(0.1)
        pyautogui.press("esc")  # Press esc to dismiss confirm dialog box

##### Key Page Configuration #####
""" Main Function to Fill in the inputs."""

class KeyPage(ctk.CTkFrame):
    def __init__(self, parent, controller):
        super().__init__(parent)
        self.controller = controller

        global pathVar, cellVar, sheetVar

        pathVar = tk.StringVar()
        cellVar = tk.StringVar()
        sheetVar = tk.StringVar()

        self.configure(fg_color="#ECEFCA")

        self.title2 = ctk.CTkLabel(self,
            text="Keying Page", 
            font=("Arial", 26, "bold"), 
            text_color="#213448",
            anchor="center",
            width=700
            )
        self.title2.place(relx=0.5, rely=0.2, anchor=CENTER)

        self.back_default_img = ImageTk.PhotoImage(Image.open("components/defaultback.png").resize((60, 60), Image.Resampling.BICUBIC))
        
        self.key_page_button = ctk.CTkButton(self, 
            image=self.back_default_img, 
            width=40, height=40,
            text="",
            fg_color="transparent",
            hover_color="#ECEFCA",
            command=lambda: controller.show_page("Home")
            )
        self.key_page_button.place(relx=0.1, rely=0.2, anchor=CENTER)

        self.file_path_entry = ctk.CTkEntry(self, 
            placeholder_text="Enter Excel File Path...", 
            width=300, 
            height=40, 
            textvariable=pathVar
            )
        self.file_path_entry.pack(padx=(160, 3), side=ctk.LEFT)

        self.browse_button = ctk.CTkButton(self, 
            text="Browse...", 
            width=100, 
            height=40, 
            command=self.browse_file
        )
        self.browse_button.pack(padx=(0, 0), side=ctk.LEFT)

        self.cell_entry = ctk.CTkEntry(self,
            placeholder_text="Enter Starting Cell (Ex. C2)", 
            width=55, 
            height=40, 
            textvariable=cellVar
            )
        self.cell_entry.place(relx=0.26, rely=0.62, anchor=CENTER)

        self.sheet_entry = ctk.CTkEntry(self,
            placeholder_text="Enter Sheet Name (Ex. Sheet1)", 
            width=100, 
            height=40, 
            textvariable=sheetVar
            )
        self.sheet_entry.place(relx=0.65, rely=0.62, anchor=CENTER)

        self.begin_keying_button = ctk.CTkButton(self,
            text="Confirm", 
            font=("Arial", 17, "bold"), 
            width=200, 
            height=40, 
            command=lambda: self.controller.show_page("countDown2")
        )
        self.begin_keying_button.place(relx=0.5, rely=0.8, anchor=CENTER)

    def browse_file(self):
        """Opens file dialog and sets the selected path into the entry."""
        file_path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel Files", "*.xlsx *.xls")]
        )
        if file_path:
            pathVar.set(file_path)

class countDownPage2(ctk.CTkFrame):
    def __init__(self, parent, controller, countdown_time=5):
        super().__init__(parent)
        self.controller = controller
        self.countdown_time = countdown_time
        self._cancelled = False  # Add cancellation flag

        self.configure(fg_color="#ECEFCA")

        self.count_text = ctk.CTkLabel(self, text="Starting in", font=("Arial", 26, "bold"), text_color="#213448")
        self.count_text.place(relx=0.5, rely=0.3, anchor=CENTER)

        self.count = ctk.CTkLabel(self, text="", font=("Arial", 40, "bold"), text_color="#213448")
        self.count.place(relx=0.5, rely=0.4, anchor=CENTER)

        # Home button (hidden at start)
        self.home_button = ctk.CTkButton(
            self,
            text="Return to Home",
            font=("Arial", 17, "bold"),
            width=200,
            height=40,
            command=lambda: self.return_home()  # Use method to return to home
        )
        self.home_button.place(relx=0.5, rely=0.6, anchor=CENTER)
        self.home_button.lower()  # Hide initially

        # Cancel button
        self.cancel_button = ctk.CTkButton(
            self,
            text="Cancel",
            font=("Arial", 17, "bold"),
            width=200,
            height=40,
            command=self.cancel_countdown  # Cancel countdown
        )
        self.cancel_button.place(relx=0.5, rely=0.7, anchor=CENTER)

    def cancel_countdown(self):
        """Cancel countdown and reset the flag."""
        self._cancelled = True  # Set the cancel flag
        global stop_flag
        stop_flag = True  # Set stop_flag to True to stop the process
        self.count.configure(text="Cancelled", text_color="red")
        cellVar.set("")  # Clear the entry field after processing
        pathVar.set("")
        sheetVar.set("")
        self.count_text.configure(text="")  # Clear the "Starting in" text
        self.home_button.lift()  # Show the Home button so user can go back manually
        self.cancel_button.lower()  # Hide the Cancel button after cancellation

    def return_home(self):
        """Return to Home and reset flags."""
        global stop_flag
        stop_flag = False  # Reset stop flag for future runs
        self._cancelled = False  # Reset cancel flag
        self.count_text.configure(text="Starting in", text_color="#213448")
        self.controller.show_page("Home")  # Go back to home

    def start_countdown(self):
        self.controller.pages["Key"].cell_entry.configure(border_color="#A9A9A9")
        self.controller.pages["Key"].sheet_entry.configure(border_color="#A9A9A9")
        self.controller.pages["Key"].file_path_entry.configure(border_color="#A9A9A9")
        self.remaining = self.countdown_time
        self.cancel_button.lift()
        self.home_button.lower()  # Hide button if revisiting
        self.count.configure(text="", text_color="#213448")
        self._update_timer()

    def _update_timer(self):
        if self.remaining >= 0 and not self._cancelled:  # Check if not cancelled
            self.count.configure(text=self.remaining)
            self.remaining -= 1
            self.update()  # Force UI update to display Cancel button
            self.after(1000, self._update_timer)
        else:
            if not self._cancelled:
                self.count_text.configure(text="")
                self.count.configure(text="Keying has started!")
                self.after(500, self.on_key_finished)  # Call on_key_finished after countdown
            else:
                # If cancelled, stop further action and reset
                self.count.configure(text="Cancelled", text_color="red")

    def on_key_finished(self):
        """Start the keying process in a separate thread."""
        thread = threading.Thread(target=self.start_key_thread)
        thread.daemon = True # Make thread a daemon so it exits when the main program exits
        thread.start()

    def start_key_thread(self):
        """Runs the clearing logic and updates UI from main thread."""
        self.start_key()

    def start_key(self):
        global cellVar, sheetVar, pathVar, stop_flag
        excelfile = pathVar.get()
        cell = cellVar.get()
        sheet = sheetVar.get()

        cellVar.set("")  # Clear the entry field after processing
        pathVar.set("")
        sheetVar.set("")

        try:
            # Step 1: Read data from the Excel column
            values = self.read_excel_column(excelfile, sheet, cell)
            if not values or values == ["", None]:
                
                return

            for value in values:
                if stop_flag:
                    print("Process stopped by user.")
                    break
                print(f"Typing value: {value}")
                self.type_to_program(value)
                time.sleep(0.6)
            
            if not stop_flag:
                self.after(1000, self._update_done)  # Update UI after process is done
            else:
                self.after(1000, self._update_cancelled)  # Update UI after process is cancelled

        except Exception as e:
            self.count.configure(text="No Data Found.", text_color="red")
            self.cancel_button.lower()  # Hide the cancel button after finishing
            self.home_button.lift()
        
    def _update_done(self):
        """Update the UI after the clearing process is done."""
        self.count.configure(text="Done!", text_color="green")
        self.cancel_button.lower()  # Hide the cancel button after finishing
        self.home_button.lift()  # Show the home button after finishing

    def _update_cancelled(self):
        """Update the UI after the clearing process is cancelled."""
        self.count.configure(text="Cancelled", text_color="red")
        self.home_button.lift()  # Show the home button after cancelling

    def read_excel_column(self, file_path, sheet_name, start_cell):
        """Read values from a column in an Excel file until an empty cell is encountered."""
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]

        # Determine the starting row and column
        start_column = openpyxl.utils.cell.column_index_from_string(start_cell[:1])  # Column letter to number
        start_row = int(start_cell[1:])  # Starting row number

        # Read values from the column
        values = []
        for row in range(start_row, sheet.max_row + 1):
            cell_value = sheet.cell(row=row, column=start_column).value
            if cell_value is None:  # Stop at the first empty cell
                break
            values.append(cell_value)

        return values

    def type_to_program(self, data):
        """Type data directly into another program."""
        # Type the data as a string
        pyautogui.typewrite(str(data))  # Utilizes pyautogui library to simulate user typing
        pyautogui.press("enter")  # Will simulate the enter key being pressed
        pyautogui.press("esc")  # Dismisses the popup confirmation after keying in a quantity


##### Home Page Configuration #####
""" Hub Page to access the different Functions. """

class HomePage(ctk.CTkFrame):
    def __init__(self, parent, controller):
        super().__init__(parent)
        self.controller = controller

        self.configure(fg_color="#ECEFCA")

        title1 = ctk.CTkLabel(self,
        text="Welcome to the PI Count AutoKey Program", 
        font=("Arial", 26, "bold"), 
        text_color="#213448", 
        anchor="center", 
        width=700
        )
        title1.place(relx=0.5, rely=0.2, anchor=CENTER)

        subtitle1 = ctk.CTkLabel(self,
        text="Are You Keying or Clearing?",
        font=("Arial", 15, "bold"),
        text_color="#213448",
        anchor="center",
        width=700
        )
        subtitle1.place(relx=0.5, rely=0.33, anchor=CENTER)

        buttons_frame = ctk.CTkFrame(self, fg_color="#ECEFCA", width=400, height=60)
        buttons_frame.place(relx=0.5, rely=0.5, anchor=CENTER)

        key_page_button = ctk.CTkButton(buttons_frame, text="KEY", font=("Arial", 17, "bold"), width=100, height=50,command=lambda: controller.show_page("Key"))
        key_page_button.pack(padx=(0,33), side=ctk.LEFT)
        
        back_button = ctk.CTkButton(buttons_frame, text="CLEAR", font=("Arial", 17, "bold"), width=100, height=50,command=lambda: controller.show_page("Clear"))
        back_button.pack(padx=(0,0), side=ctk.RIGHT)

##### Program Set up #####
""" Sets up Window for Program, keep track of the pages for navigation. """

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("PI Count AutoKey")
        self.geometry("720x480")

        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)

        # Container for all pages
        container = ctk.CTkFrame(self)
        container.pack(fill="both", expand=True)

        # Let container expand
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        # Dictionary to hold references to pages
        self.pages = {}

       # Create and grid pages, hide all initially
        for PageClass, name in [(HomePage, "Home"), 
            (KeyPage, "Key"), 
            (ClearPage, "Clear"), 
            (countDownPage, "countDown"),
            (countDownPage2, "countDown2")
            ]:
            page = PageClass(container, self)
            self.pages[name] = page
            page.grid(row=0, column=0, sticky="nsew")
            page.lower()  # Hide by default

        self.show_page("Home")  # Show the home page first

    def show_page(self, page_name):
        """Show a page by name."""
        for page in self.pages.values():
            page.lower()

        # Raise only the one we want
        self.pages[page_name].tkraise()

        # Start Countdown if on the countdown page
        if page_name == "countDown":
            self.pages[page_name].start_countdown()
        
        if page_name == "countDown2":
            self.pages[page_name].start_countdown()

if __name__ == "__main__":
    app = App()
    app.mainloop()
